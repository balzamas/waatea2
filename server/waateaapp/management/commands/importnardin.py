from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from datetime import datetime, timedelta
from waateaapp.models import HistoricalGame
from waatea_2.users.models import UserProfile# Import your models
from django.db import IntegrityError
from pathlib import Path


class Command(BaseCommand):
    help = 'Import historical game data from Nardins database'

    #Column 0: Date
    #Column 1: Winti Team
    #Column 2: Opposition Team
    #Column 3: Competition
    #Column 4 - 26: Players Winti (License name)


    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Path to the Excel file')

    def handle(self, *args, **kwargs):

        file_path = kwargs['file_path']

        workbook = load_workbook(filename=str(file_path))
        sheet = workbook.active

        row_count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            print ("importing: " + row[1] + " - " + row[2])
            date = row[0] + timedelta(hours=12)
            played_for = row[1].encode('latin-1').decode('utf-8')
            played_against = row[2].encode('latin-1').decode('utf-8')
            competition = row[3].encode('latin-1').decode('utf-8')

            position = 0
            col = 0

            for cell_value in row:
                col = col+1
                if col > 4 and cell_value != None:
                    position = position + 1

                    try:
                        player = UserProfile.objects.get(sportlomo_id=cell_value).user

                        HistoricalGame.objects.create(
                            played_for=played_for,
                            played_against=played_against,
                            player=player,
                            date=date,
                            position=position,
                            competition=competition,
                        )

                        self.stdout.write(self.style.SUCCESS(f'Successfully imported game for {cell_value}'))
                    except UserProfile.DoesNotExist:
                        self.stderr.write(self.style.ERROR(f'Player {cell_value} not found'))
                    except IntegrityError as e:
                        self.stderr.write(self.style.ERROR(f'Record exists'))
                    except:
                        self.stderr.write(self.style.ERROR(f'Something failed'))


                    row_count = row_count + 1

            self.stdout.write(self.style.SUCCESS('Import complete'))

