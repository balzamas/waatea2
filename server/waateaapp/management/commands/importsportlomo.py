from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from datetime import datetime
from waateaapp.models import HistoricalGame
from waatea_2.users.models import UserProfile# Import your models
from django.db import IntegrityError
from pathlib import Path

class Command(BaseCommand):
    help = 'Import historical game data from Excel files - Sportlomo CH Edition (*.xlsx)'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Path to the Excel files')

    def handle(self, *args, **kwargs):

        file_path = kwargs['file_path']

        files = Path(file_path).glob('*.xlsx')
        for filename in files:
            print(filename)
            print(filename)
            workbook = load_workbook(filename=file_path + "/" + str(filename))
            sheet = workbook.active

            sportlomo_id = sheet.cell(row=1, column=3).value

            row_count = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row_count > 4:
                    position = row[0]

                    if position != 'Team Official':
                        date_str = row[6]

                        try:
                            player = UserProfile.objects.get(sportlomo_id=sportlomo_id).user

                            played_for = row[4]
                            played_against = row[5]
                            print(date_str)
                            date = datetime.strptime(date_str, '%d/%m/%Y %H:%M')  # Adjust format as needed
                            competition = row[2]

                            HistoricalGame.objects.create(
                                played_for=played_for,
                                played_against=played_against,
                                player=player,
                                date=date,
                                position=position,
                                competition=competition,
                            )

                            self.stdout.write(self.style.SUCCESS(f'Successfully imported game for {sportlomo_id}'))
                        except UserProfile.DoesNotExist:
                            self.stderr.write(self.style.ERROR(f'Player {sportlomo_id} not found'))
                        except IntegrityError as e:
                            self.stderr.write(self.style.ERROR(f'Record exists'))

                row_count = row_count + 1

        self.stdout.write(self.style.SUCCESS('Import complete'))
